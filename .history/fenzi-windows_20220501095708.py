# -*- encoding:utf-8 -*-
from ast import Expression
import os
import pandas as pd
from tkinter.tix import COLUMN
import openpyxl
import shutil
import datetime
from ctypes import alignment
from openpyxl import Workbook
from openpyxl  import load_workbook #用于读取Excel中的信息
from openpyxl.styles import Font,colors,Alignment
from openpyxl.styles import PatternFill #单元格填充

nowTime = datetime.datetime.now().strftime('%Y-%m-%d') # 日期
nowTime_m = datetime.datetime.now().strftime('%H-%M') # 到小时分
data_list = {}
rv_col = ["[Instrument]","仪器型号","仪器编号"]  


# 定义函数转换.xls文件为xlsx文件
def Chang_file_type(file):
    df = pd.read_excel(file,header=None)
    df.to_excel(root + "\\"+name+"x", index=False,header=False) #不写入索引列与windows
    # df.to_excel(root + "/"+name+"x", index=False,header=False) #mac
    file = os.path.join(root,name+"x") 


#定义函数对文件进行排序
def Chang_file_sort(fargs,*args):
    args = list(args) #元组转换为列表
    df = pd.read_excel(fargs,header=None)
    df.sort_values(args,inplace=True ) #默认从小到大

#定义函数设置文件格式
def Set_styl (fargs,*args):
    wb = load_workbook(fargs)
    ws = wb.active 
    for row in range(1, ws.max_row+1):
        for col in range(1,9):
            c = ws.cell(row=row,column=col)
            c.font = Font(size=16)
    ws.column_dimensions['H'].width = 20 #调整列宽
    ws.column_dimensions["B"].width = 12 #调整列宽
    font = Font(size= 16, bold=True, italic=False, strike=False, color='FF0000') # 设置字体样式    
    for i in range(1,ws.max_row+1):
        ws.row_dimensions[i].height = 30 #设置行高
        if args[0] in str(ws["C"+str(i)].value):
            ws["C"+str(i)].font = font
        if args[1] in str(ws["H"+str(i)].value):
            ws["H"+str(i)].font = font
    wb.save(fargs)        

    
#定义函数，7500自动编入样本号
def Fill_system_7500(file):
    global first,contr_sum,contr_n
    list = ("A","B", "C","D","E","F","G","H")
    wb = load_workbook(file)
    ws = wb.active    
    contr_n = []
    start_n = first #开始计数号
    first_col = start_n #单板总计数
    
    for n in range(1,13):
        first_n = first_col                   
        for col in list:
            for row in range(initial_row+1,ws.max_row+1):
                Well_A = ws["A"+str(row)].value #取出A列值
                Well_B = ws["B"+str(row)].value #取出B列的值,是否为待测样品
                Well_B = Well_B[0:1]
                contr_list = ["N","P"] #质控列表
                contr_sum = len(contr_n) 
                if Well_A == str(col+str(n)):
                    if Well_B in contr_list :
                        pass # 保留原B列质控信息
                        # ws.cell(row = row,column=2,value="") #B列质控编号置空
                        if Well_A not in contr_n:
                            contr_n.append(Well_A)
                        continue
                    ws["B"+str(row)] = str(int(first_n)- int(contr_sum)) #B列的编号
            first_n = int(first_n) +1 
        first_col = first_n 
    wb.save(file)

#定义函数，HS自动编入样本号
def Fill_system_hs(file):
    global first,contr_sum,contr_n
    list = ("A","B", "C","D","E","F","G","H")
    wb = load_workbook(file)
    ws = wb.active    
    contr_n = []
    start_n = first #开始计数号
    first_col = start_n #单板总计数
    contr_list = ["阴性对照","阳性对照"] #质控列表
    for n in range(1,13):
        first_n = first_col                   
        for col in list:
            for row in range(initial_row+1,ws.max_row+1):
                Well_A = ws["A"+str(row)].value  # 取出A列值
                Well_J = ws["J"+str(row)].value  # 取出J列的值,是否为待测样品
                if Well_J in contr_list:
                    ws.cell(row = row,column=25,value="") # 给Y列赋值为空（质控）
                if Well_A == str(col+str(n)):
                    if Well_J in contr_list :
                        if Well_A not in contr_n:
                            contr_n.append(Well_A) #增加质孔位
                        continue
                    contr_sum = len(contr_n) # 计算质控数量
                    ws["Y"+str(row)] = str(int(first_n)- int(contr_sum)) #Y列的编号
            first_n = int(first_n) +1 
        first_col = first_n 
    wb.save(file)
    
#定义函数判断结果
def Get_system_result(fargs):
    global data_list
    Contr_edit(fargs)
    data_list = dict(A1="板位",B1="样本号",C1="结果",D1="ORF1ab",E1="N",F1="IC",G1="E",H1="建议")
    wb = load_workbook(fargs)
    ws = wb.active     
    acol_value = set() # A列集合数
    for row in range(initial_row+1,ws.max_row+1):
        acol_value.add(ws["A"+str(row)].value)
    n=2
    for A_value in acol_value:
        for row in range(initial_row+1,ws.max_row+1):
            Well_cont = ws[sample_type+str(row)].value # 样本类型列的值，主要是用来区分质控样本
            if Well_cont in contr_list and A_value == ws["A"+str(row)].value:
                n = n+1
                qc_n = n//3
                ws[num_col+str(row)] = Well_cont+str(qc_n) # 样本号列赋值质控编号
    
    
    
    hole_list =[]
    sample_num = []
    for row in range(initial_row+1,ws.max_row+1):
        sample_N = ws[num_col+str(row)].value #取出样本号
        if sample_N not in sample_num:
            sample_num.append(sample_N)
    list = ("A","B", "C","D","E","F","G","H",)    
    for i in range(1,13):
        for n in list:
            hole = (str(n)+str(i))
            hole_list.append(hole) # 产生A1～H12的列表清单
    row_total = int(1)
    for sample_n in sample_num: #增加核酸项
        Result = "阴性"
        Gen_ct_list ={} #ct值字典
        for row in range(initial_row+1,ws.max_row+1):
            data_A = ws["A"+str(row)].value #获取A列值
            Gen_name = ws[gen_col+str(row)].value #获基因名列中的具体基因
            ct_value = ws[ct_col+str(row)].value #获取CT值          
            sample_value = ws[num_col+str(row)].value # 获取样本号
            if Gen_name not in("ORF1ab","N","E","IC"):
                continue
            if sample_n == sample_value and Gen_name == "ORF1ab":
                if  ct_value == NoCt:
                    ORF_v = NoCt
                    Gen_ct_list['ORF_v'] = ORF_v
                else:
                    ORF_v = ws[ct_col+str(row)].value #取出ORF1ab值
                    Gen_ct_list['ORF_v'] = ORF_v
            if sample_n == sample_value and Gen_name == "N" :
                if ct_value == NoCt:
                    N_v = NoCt
                    Gen_ct_list['N_v'] = N_v
                else:
                    N_v = ws[ct_col+str(row)].value #取出N基因值
                    Gen_ct_list['N_v'] = N_v
            if sample_n == sample_value and Gen_name == "E" : 
                if ct_value == NoCt:
                    E_v = NoCt
                    Gen_ct_list['E_v'] = E_v
                else:
                    E_v = ws[ct_col+str(row)].value #取出E基因值
                    Gen_ct_list['E_v'] = E_v      
            if sample_n == sample_value and Gen_name == "IC" :  
                if ct_value == NoCt:
                    IC_v = NoCt
                    Gen_ct_list['IC_v'] = IC_v
                else:
                    IC_v = ws[ct_col+str(row)].value #取出IC值
                    Gen_ct_list['IC_v'] = IC_v                              
            if sample_n == sample_value:
                cell_N = data_A
        if "ORF_v" not in Gen_ct_list.keys():
            continue
        if IC_v == NoCt: #IC不起要求复检
            try :
                int(float(ORF_v))
                int(float(N_v))
                if float(N_v)<= int(35) and float(ORF_v) <= int(35): # 需要核查原始数据
                    result = "阳性"
                    advise = "复检IC未起,双靶阳性"
            except:
                result = "IC未起"
                advise = "复检,IC未起"
        elif ORF_v == NoCt and N_v == NoCt: #IC起，同时两基因NoCt为阴性不需复查，直报
            result = "阴性"
            advise = "报送" 
        elif ORF_v == NoCt and float(N_v) > int(35): #IC起 ORF不起,同时N大于35判为阴性，直报
            result = "阴性"
            advise = "报送" 
        elif ORF_v == NoCt and float(N_v) <= int(35): #IC起，ORF不起，单N阳性，复检
            result = "阳性"
            advise = "复检,N单靶阳" 
        elif N_v == NoCt  and float(ORF_v) <= int(35): #IC起，单ORF阳，N不起，复检
            result = "阳性"
            advise = "复检,ORF单靶阳" 
        elif N_v == NoCt and float(ORF_v)>int(35): # N不起，ORF阴性（符合条件阴性，单基因大于35,直报
            result = "阴性" 
            advise = "报送"   
        elif float(N_v)>= int(35) and float(ORF_v) < int(35): #均有CT值，N大于域值，ORF单阳，复检
            result = "阳性"
            advise = "复检,ORF单靶阳" 
        elif float(N_v) < int(35) and float(ORF_v) >= int(35): # #均有CT值，ORF大于域值，N单阳，复检
            result = "阳性"
            advise = "复检,N单靶阳" 
        elif  int(35) <= float(N_v) < int(40) and int(35)<= float(ORF_v) < int(40): # #均有CT值，ORF，N值均在可疑区间，复检
            result = "阳性"
            advise = "双靶灰区,复检" 
        elif int(40) <= float(N_v) and  int(40)<= float(ORF_v): # ORF，N均大于域值（38）
            result = "阴性"
            advise = "报送" 
        elif float(N_v)< int(35) and   float(ORF_v)<int(35): # ORF，N均小于域值（35）
            result = "阳性"
            advise = "报送" 
        elif int(35) <= float(N_v)<int(40) and  int(40)<= float(ORF_v): # ORF，N均大于域值（38）
            result = "阴性"
            advise = "报送"         
        elif int(40) <= float(N_v) and  int(35)<= float(ORF_v)<int(40): # ORF，N均大于域值（38）
            result = "阴性"
            advise = "报送"              
        print(sample_n,Gen_ct_list,result,advise)
        data_list[str("A")+str(row_total+1)] = str(cell_N)
        data_list[str("B")+str(row_total+1)] = str(sample_n)
        data_list[str("C")+str(row_total+1)] = str(result)
        data_list[str("D")+str(row_total+1)] = str(ORF_v)
        data_list[str("E")+str(row_total+1)] = str(N_v)
        data_list[str("F")+str(row_total+1)] = str(IC_v)
        data_list[str("H")+str(row_total+1)] = str(advise)
        if "E_v" in Gen_ct_list.keys():
            data_list[str("G")+str(row_total+1)] = str(E_v)
        row_total = row_total+1      
    return(data_list)    
    '''    
        # 在原表后续行中加入判断结果
        data_list[str("A")+str(row_total+1)] = hole_n
        data_list[str(gen_col)+str(row_total+1)] = "Result" #在基因项目名称列增加核酸结果项
        ORF_v = Gen_ct_list['ORF_v']
        N_v = Gen_ct_list['N_v']
        E_v = Gen_ct_list['E_v']
        IC_v = Gen_ct_list["IC_v"]
        if int(float(ORF_v)) <= int(35) or int(float(N_v)) <= int(35): #结果判断规则
            Result = "阳性"     
        data_list[str(ct_col)+str(row_total+1)] = Result #在对应CT列增加核酸CT值
        data_list[str(num_col)+str(row_total+1)] = sample_N #样本号列加入样本号
    wb.save(file)
    
    wb = Workbook() #新建文件
    ws = wb.active
    '''

#定义函数达安试剂CT加10
gen_list = ["N","E","ORF1ab"]
def Daan_ct(file):
    wb = load_workbook(file)
    ws = wb.active 
    for row in range(initial_row,ws.max_row+1):
        Gen_name = ws[gen_col+str(row)].value 
        ct_value = ws[ct_col+str(row)].value 
        if Gen_name in gen_list and ct_value != NoCt:
            ct_value = str(round((float(ct_value)+10),2))
            ws[ct_col+str(row)] = ct_value
    wb.save(file)

# 定义函数，对质控位进行编号
def Contr_edit(fargs):
    wb = load_workbook(fargs)
    ws = wb.active 
    global contr_list
    contr_n =[]
    acol_value = set() # A列集合数
    for row in range(initial_row+1,ws.max_row+1):
        acol_value.add(ws["A"+str(row)].value)
    n=2
    for A_value in acol_value:
        for row in range(initial_row+1,ws.max_row+1):
            Well_cont = ws[sample_type+str(row)].value # 样本类型列的值，主要是用来区分质控样本
            if Well_cont in contr_list and A_value == ws["A"+str(row)].value:
                n = n+1
                qc_n = n//3
                ws[num_col+str(row)] = Well_cont+str(qc_n) # 样本号列赋值质控编号
    # wb.save(fargs)

#定义函数，转移文件    
def move_file(work_path,new_path,file_typ,args):
    if not os.path.exists(new_path): #如果文件夹不存在时进行创建
        os.makedirs(new_path)
    if args==0 : # 对等待处理的文件进行提前备份或移除
        for derName, subfolders, filenames in os.walk(work_path):
            for i in range(len(filenames)):
                if filenames[i].endswith(file_typ):
                    file_path=derName+'\\'+filenames[i] #windows
                    newpath=new_path+'\\'+filenames[i] #windows
                    # file_path=derName+'/'+filenames[i] #mac
                    # newpath=new_path+'/'+filenames[i] #mac
                    shutil.move(file_path,newpath)
    if args == 1: # 对新生成的，已处理结束的文件进行备份
        for filenames in os.listdir(work_path):
            if filenames.endswith(file_typ) and not name.startswith("$") and not name.startswith(".")and not name.startswith("~"):
                file_path=work_path+'\\'+filenames #windows
                newpath=new_path+'\\'+filenames[0:-5]+nowTime_m+file_typ #windows
                # file_path=work_path+'/'+filenames #mac
                # newpath=new_path+'/'+filenames[0:-5]+nowTime_m+file_typ #mac
                shutil.move(file_path,newpath)
    if args == 2: # 单个文件转移
        shutil.move(work_path,new_path)

#定义函数，拷贝文件    
def copy_file(work_path,new_path,file_typ,args): # args:0所有文件含子文件夹下的文件，args：1是当前文件夹，args:2明确的单个文件
    if not os.path.exists(new_path): #如果文件夹不存在时进行创建
        os.makedirs(new_path)
    if args==0 :
        for derName, subfolders, filenames in os.walk(work_path):
            for i in range(len(filenames)):
                if filenames[i].endswith(file_typ):
                    file_path=derName+'\\'+filenames[i] #windows
                    newpath=new_path+'\\'+filenames[i] #windows
                    # file_path=derName+'/'+filenames[i] #mac
                    # newpath=new_path+'/'+filenames[i] #mac
                    shutil.copyfile(file_path,newpath)
    if args == 1: # 对新生成的，已处理结束的文件进行备份
        for filenames in os.listdir(work_path):
            if filenames.endswith(file_typ) and not name.startswith("$") and not name.startswith(".")and not name.startswith("~"):
                file_path=work_path+'\\'+filenames #windows
                newpath=new_path+'\\'+filenames[0:-5]+"S"+nowTime_m+file_typ #windows
                # file_path=work_path+'/'+filenames #mac
                # newpath=new_path+'/'+filenames[0:-5]+"S"+nowTime_m+file_typ #mac
                shutil.copyfile(file_path,newpath)
    if args == 2: # 单个文件转移
        shutil.copy(work_path,new_path+file_typ)


#定义函数获取列值 ,并写入汇总表 
def Get_system_data(fargs):
    global row_total,data_list 
    wb = load_workbook(file)
    ws = wb.active
    #data_need_idx = list(enumerate(data_need)) #对列表进行编号
    for row in range(initial_row+1,ws.max_row+1):
        for col in data_need:
            data = ws[col+str(row)].value #不同列信息
            data_A = ws["A"+str(row)].value
            if data_A == None:#过滤空值
                return
            if data_A in rv_col:#过滤无用的表尾值
                return
            data_list[str(col)+str(row_total+1)] = data
        row_total = row_total + 1  #下一个个文件的开始行               
    # os.remove(fargs)
    return (data_list,row_total)

# 新建工作表
def Creat_file(fargs,**kwargs):
    wb = Workbook() #新建文件
    ws = wb.active    
    for key,item in  kwargs.items(): #写入文件
        ws[key]=item
    wb.save(fargs)

# 修改域值函数
def Threa_edit(low_thread,high_thread):
    if flag == false:
        low_thread = input("低域调整为:\n\n")
        high_thread = input("高域调整为:\n\n")
    low_thread = low_thread
    high_thread = high_thread
if __name__ == '__main__':
    while True:
        choice_f = input('处理7500还是宏石文件:\n\n1.7500\n\n2.宏石\n\n')
        if choice_f == "1":
            work_path = os.getcwd() + "\\7500" # 获取当前工作路径，并指定文件夹 windows
            finial_name = "\\7500.xlsx"
            finial_file = os.getcwd()+finial_name
            result_name = "\\reault_7500.xlsx"
            result_file = os.getcwd()+result_name
            contr_list = ["N","P"]
            # work_path = os.getcwd() + "/test" # 获取当前工作路径，并指定文件夹 mac
            # finial_file = os.getcwd()+"/7500.xlsx"   # mac
            # result_file = os.getcwd()+"/reault_7500.xlsx"
            data_need = ["A","B","C","E","G"] #需要获取的信息内容列
            initial_row = 8
            row_total = initial_row #宏石-14,7500-8 
            name_must =dict(A8="Well",B8="Sample Name",C8="Target Name",G8="Ct")
            NoCt = "Undetermined" #无Ct值时的赋值
            gen_col = "C" #基因列
            ct_col = "G" #Ct值列
            num_col = "B" #样本号列
            sample_type = "B" # 样本类型列
            back_path = os.getcwd()+"\\back\\7500back\\"+str(nowTime) #windows
            # back_path = os.getcwd()+"/back/7500back/"+str(nowTime) #mac
            for root,dirs,files in os.walk(work_path) :  
                for file in files:
                    name = file
                    file = os.path.join(root,file)    #
                    if name.endswith(".xls") and not name.startswith("$") and not name.startswith(".")and not name.startswith("~"):
                        Chang_file_type(file) #转变文件类型
            move_file(work_path,back_path,".xls",0) #备份原文件
        if choice_f == "2": # 宏石
            work_path = os.getcwd() + "\\HS" #获取当前工作路径，并指定文件夹 windows
            finial_name = "\\HS.xlsx"
            finial_file = os.getcwd()+finial_name
            result_name = "\\reault_HS.xlsx"
            result_file = os.getcwd()+result_name
            contr_list = ["阴性对照","阳性对照"]
            # result_file = os.getcwd()+"/reault_HS.xlsx" #判断结果最后的文件 mac
            # work_path = os.getcwd() + "/test" #获取当前工作路径，并指定文件夹 mac
            # finial_file = os.getcwd()+"/HS.xlsx" #mac
            name_must =dict(A14="反应孔",F14="通道",G14="染料",H14="目标",J14="类型",M14="Ct",Y14="唯一标识")
            initial_row = 14 
            row_total = initial_row #宏石-14,7500-8 
            NoCt = "NoCt"  #无Ct值时的赋值
            gen_col = "H" #基因列
            ct_col = "M"  #Ct值列
            num_col = "Y" #样本号列
            sample_type = "J" #样本类型列
            data_need = ["A","F","G",gen_col,sample_type, ct_col,num_col] #需要获取的信息内容列
            back_path = os.getcwd()+"\\back\\HSback\\"+str(nowTime) #windows
            # back_path = os.getcwd()+"/back/HSback/"+str(nowTime) #mac
            copy_file(work_path,back_path,".xlsx",0) #按文件夹查找（含子文件夹）备份原文件
            # move_file(work_path,back_path,".xlsx") #移走原文件
        choice_edit = input('如何对板孔进行编号处理：\n\n1.已有样本编号只需对CT值加十处理\n\n2.需要编号并对CT加十处理\n\n3.仅需对板子进行编号\n\n4.仅判断结果，不做其他处理\n\n,5.直接退出\n\n')
        if choice_edit == "5":
            break
        for root,dirs,files in os.walk(work_path) :  
            for file in files:
                name = file
                file = os.path.join(root,file)            
                if name.endswith(".xlsx") and not name.startswith("$") and not name.startswith(".")and not name.startswith("~"):
                    if choice_edit == "1" :
                        #print(file)
                        Daan_ct(file) 
                    if choice_edit == "2" :
                        print(file)
                        first = int(input("本板样本起始号：")) #输入板子起始样本号
                        Daan_ct(file)    
                        if choice_f == "1":
                            Fill_system_7500(file)
                        if choice_f == "2":
                            Fill_system_hs(file)    
                    if choice_edit == "3" :
                        print(file)
                        first = int(input("本板样本起始号：")) #输入板子起始样本号
                        if choice_f == "1":
                            Fill_system_7500(file)
                        if choice_f == "2":
                            Fill_system_hs(file) 
                    if choice_edit == "4" :
                        pass
                    Get_system_data(file)
                    os.remove(file)
        Creat_file(finial_file,**name_must,**data_list) #生成汇总文件
        # Contr_edit(finial_file)
        Get_system_result(finial_file)  # 对汇总的文件判定结果
        Creat_file(result_file,**data_list) # 创建复查文件
        Set_styl(result_file,*["阳性"],*["复检"])
        back_path_finial = back_path+"\\板条处理完成文件" # windows
        back_path_result = back_path+"\\结果判读文件" # windows
        # back_path_finial = back_path+"/板条处理完成文件" # mac
        # back_path_result = back_path+"/结果判读文件" # mac
        file_typ = finial_name[0:-5]+"S"+nowTime_m+".xlsx" 
        copy_file(finial_file,back_path_finial,file_typ,2) #保存处理好的文件
        file_typ = result_name[0:-5]+"S"+nowTime_m+".xlsx" 
        copy_file(result_file,back_path_result,file_typ,2) #保存处理好的文件 
        
        print("已完成")
        os.system("pause")
        break

